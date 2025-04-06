from PyQt5.QtWidgets import (QDialog, QFormLayout, QLineEdit, QPushButton,
                             QHBoxLayout, QWidget, QRadioButton, QButtonGroup,
                             QComboBox, QMessageBox, QLabel, QCompleter)
from PyQt5.QtCore import Qt, QStringListModel
from openpyxl import Workbook, load_workbook
import os


class NewCaseWindow(QDialog):
    IDENTITY_MAP = {
        1: "本人",
        2: "证人",
        3: "法人"
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建工伤案件")
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinMaxButtonsHint)
        self.resize(600, 500)
        self.excel_file = "company_data.xlsx"
        self._init_excel()  # 初始化 Excel 文件（改用 openpyxl）
        self._init_ui()
        self._connect_identity_signals()

    def _init_excel(self):
        """初始化Excel数据文件（竖向排列）"""
        if not os.path.exists(self.excel_file):
            # 创建新的Excel文件，数据竖向排列
            wb = Workbook()

            # 公司名称（A列）
            ws_company = wb.active
            ws_company.title = "公司名称"
            ws_company.append(["示例公司A"])  # A1
            ws_company.append(["示例公司B"])  # A2

            # 用人单位（B列）
            ws_employer = wb.create_sheet("用人单位")
            ws_employer.append(["部门A"])  # A1
            ws_employer.append(["部门B"])  # A2

            # 派驻单位（C列）
            ws_dispatched = wb.create_sheet("派驻单位")
            ws_dispatched.append(["分部1"])  # A1
            ws_dispatched.append(["分部2"])  # A2

            wb.save(self.excel_file)

        # 加载数据（竖向读取）
        self.wb = load_workbook(self.excel_file)

        # 公司名称（A列所有非空单元格）
        self.company_list = [
            row[0].value for row in self.wb["公司名称"].iter_rows(min_col=1, max_col=1)
            if row[0].value is not None
        ]

        # 用人单位（A列所有非空单元格）
        self.employer_list = [
            row[0].value for row in self.wb["用人单位"].iter_rows(min_col=1, max_col=1)
            if row[0].value is not None
        ]

        # 派驻单位（A列所有非空单元格）
        self.dispatched_list = [
            row[0].value for row in self.wb["派驻单位"].iter_rows(min_col=1, max_col=1)
            if row[0].value is not None
        ]

    def _connect_identity_signals(self):
        """连接身份切换的信号"""
        self.identity_self.toggled.connect(self._validate_identity_switch)
        self.identity_witness.toggled.connect(self._validate_identity_switch)
        self.identity_legal.toggled.connect(self._validate_identity_switch)

    def _validate_identity_switch(self, checked):
        """
        验证身份切换
        :param checked: 按钮是否被选中
        """
        # 只有从"本人"切出时才需要验证
        if (self.identity_witness.isChecked() or self.identity_legal.isChecked()) and checked:
            if not self.name_edit.text().strip():
                QMessageBox.warning(self, "提示", "请先填写受伤职工本人姓名")

                # 阻止切换：强制切回"本人"
                self.identity_self.setChecked(True)
                return False
        return True

    def _init_ui(self):
        layout = QFormLayout()
        self.setLayout(layout)

        # 1. 案件状态选择
        self.case_status = QComboBox(self)
        self.case_status.addItems(["正常案件", "个人案件", "死亡案件"])
        layout.addRow("案件状态:", self.case_status)

        # 2. 身份类型单选组
        self.identity_group = QButtonGroup(self)
        hbox = QHBoxLayout()

        self.identity_self = QRadioButton("本人", self)
        self.identity_witness = QRadioButton("证人", self)
        self.identity_legal = QRadioButton("法人", self)

        self.identity_group.addButton(self.identity_self, 1)
        self.identity_group.addButton(self.identity_witness, 2)
        self.identity_group.addButton(self.identity_legal, 3)
        self.identity_self.setChecked(True)

        hbox.addWidget(self.identity_self)
        hbox.addWidget(self.identity_witness)
        hbox.addWidget(self.identity_legal)
        hbox.addStretch()

        identity_widget = QWidget(self)
        identity_widget.setLayout(hbox)
        layout.addRow("身份类型:", identity_widget)

        # 3. 其他字段（全部指定父对象）
        self.name_edit = QLineEdit(self)
        self.id_edit = QLineEdit(self)
        self.phone_edit = QLineEdit(self)
        self.position_edit = QLineEdit(self)

        name_label = QLabel("姓  名:")  # 使用中文全角空格(U+3000)
        name_label.setStyleSheet("font-family: 'Microsoft YaHei';")
        layout.addRow(name_label, self.name_edit)
        layout.addRow("身份证号:", self.id_edit)
        layout.addRow("联系电话:", self.phone_edit)
        layout.addRow("工作岗位:", self.position_edit)

        self.law_combo = QComboBox(self)
        self.law_combo.setEditable(False)
        laws = [
            "条例1：",
            "条例2：",
            "条例3：",
            "条例4：",
            "条例5：",
            "条例6："
        ]
        self.law_combo.addItems(laws)
        layout.addRow("适用条例:", self.law_combo)

        # 公司名称
        self.company_edit = QLineEdit(self)
        self._setup_completer(self.company_edit, self.company_list)
        btn_save_company = QPushButton("保存", self)
        btn_save_company.clicked.connect(lambda: self._save_company_data("公司名称"))
        layout.addRow("公司名称:", self._create_input_with_button(self.company_edit, btn_save_company))

        # 用人单位
        self.employer_edit = QLineEdit(self)
        self._setup_completer(self.employer_edit, self.employer_list)
        btn_save_employer = QPushButton("保存", self)
        btn_save_employer.clicked.connect(lambda: self._save_company_data("用人单位"))
        layout.addRow("用人单位:", self._create_input_with_button(self.employer_edit, btn_save_employer))

        # 派驻单位
        self.dispatched_edit = QLineEdit(self)
        self._setup_completer(self.dispatched_edit, self.dispatched_list)
        btn_save_dispatched = QPushButton("保存", self)
        btn_save_dispatched.clicked.connect(lambda: self._save_company_data("派驻单位"))
        layout.addRow("派驻单位:", self._create_input_with_button(self.dispatched_edit, btn_save_dispatched))

        # 加载首行数据
        self._load_first_row_data()

        # 4. 提交按钮
        submit_btn = QPushButton("提交", self)
        submit_btn.clicked.connect(self._on_submit)
        layout.addRow(submit_btn)

    def _on_submit(self):
        """获取并验证所有字段数据"""
        identity = self.IDENTITY_MAP.get(self.identity_group.checkedId(), "本人")

        data = {
            "status": self.case_status.currentText(),
            "identity": identity,
            "name": self.name_edit.text().strip(),
            "id": self.id_edit.text().strip(),
            "phone": self.phone_edit.text().strip(),
            "position": self.position_edit.text().strip()
        }

        # 简单验证示例
        if not data["name"]:
            QMessageBox.warning(self, "错误", "姓名不能为空")
            return

        print("案件数据:", data)
        self.accept()

    def _create_input_with_button(self, line_edit, button):
        """创建带按钮的水平布局"""
        widget = QWidget(self)
        hbox = QHBoxLayout(widget)
        hbox.setContentsMargins(0, 0, 0, 0)
        hbox.addWidget(line_edit)
        hbox.addWidget(button)
        return widget

    def _setup_completer(self, line_edit, item_list):
        """设置输入框的自动补全"""
        completer = QCompleter(item_list, self)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        line_edit.setCompleter(completer)

    def _load_first_row_data(self):
        """加载Excel第一行数据到输入框（竖向读取）"""
        # 公司名称（A1）
        company_sheet = self.wb["公司名称"]
        first_company = company_sheet["A1"].value if company_sheet["A1"].value else ""
        self.company_edit.setText(str(first_company))

        # 用人单位（A1）
        employer_sheet = self.wb["用人单位"]
        first_employer = employer_sheet["A1"].value if employer_sheet["A1"].value else ""
        self.employer_edit.setText(str(first_employer))

        # 派驻单位（A1）
        dispatched_sheet = self.wb["派驻单位"]
        first_dispatched = dispatched_sheet["A1"].value if dispatched_sheet["A1"].value else ""
        self.dispatched_edit.setText(str(first_dispatched))


    def _save_company_data(self, column_name):
        if column_name == "公司名称":
            value = self.company_edit.text().strip()
            sheet = self.wb["公司名称"]
            data_list = self.company_list
        elif column_name == "用人单位":
            value = self.employer_edit.text().strip()
            sheet = self.wb["用人单位"]
            data_list = self.employer_list
        else:
            value = self.dispatched_edit.text().strip()
            sheet = self.wb["派驻单位"]
            data_list = self.dispatched_list

        if not value:
            QMessageBox.warning(self, "错误", f"{column_name}不能为空！")
            return

        if value in data_list:
            QMessageBox.information(self, "提示", f"该{column_name}已存在！")
            return

        try:
            # 竖向追加数据（直接 append 会自动换行）
            sheet.append([value])
            self.wb.save(self.excel_file)

            # 更新内存中的数据列表
            data_list.append(value)
            QMessageBox.information(self, "成功", f"{column_name}已保存！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {str(e)}")